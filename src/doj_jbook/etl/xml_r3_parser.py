from __future__ import annotations

import os
from typing import List, Dict, Optional, Iterable

from lxml import etree
from openpyxl import load_workbook


def _parse_float_safe(v: Optional[str]) -> Optional[float]:
    if v is None:
        return None
    try:
        # Normalize commas and handle blanks
        return float(v.replace(",", "").strip())
    except Exception:
        return None


def _norm(s: Optional[str]) -> str:
    return (s or "").strip()


def _parse_float_safe_any(v: object) -> Optional[float]:
    if v is None:
        return None
    try:
        if isinstance(v, (int, float)):
            return float(v)
        s = str(v)
        if not s.strip():
            return None
        return float(s.replace(",", "").strip())
    except Exception:
        return None


def _parse_xml_file(path: str) -> List[Dict]:
    results: List[Dict] = []
    try:
        tree = etree.parse(path)
    except Exception:
        return results
    root = tree.getroot()

    # Heuristic XPaths (replace once schema is known)
    for pe in root.xpath("//ProgramElement | //PROGRAM_ELEMENT | //Pe | //PE"):
        pe_number = pe.get("number") or pe.findtext("Number") or pe.findtext("PROGRAM_ELEMENT_NUMBER") or pe.findtext("PeNumber")
        pe_name = pe.get("name") or pe.findtext("Name") or pe.findtext("PROGRAM_ELEMENT_TITLE") or pe.findtext("PeName")
        if not pe_number:
            continue
        for proj in pe.xpath(".//Project | .//PROJECT | .//Proj"):
            project_number = proj.get("number") or proj.findtext("Number") or proj.findtext("PROJECT_NUMBER")
            project_name = proj.get("name") or proj.findtext("Name") or proj.findtext("PROJECT_TITLE")
            for cc in proj.xpath(".//CostCategory | .//COST_CATEGORY | .//Line | .//COST_LINE | .//CostItem"):
                cc_name = cc.get("name") or cc.findtext("Name") or cc.findtext("COST_CATEGORY") or cc.findtext("LINE_NAME")
                fy23 = _parse_float_safe(cc.get("FY2023") or cc.findtext("FY2023"))
                fy24 = _parse_float_safe(cc.get("FY2024") or cc.findtext("FY2024"))
                fy25_base = _parse_float_safe(
                    cc.get("FY2025Base") or cc.findtext("FY2025Base") or cc.findtext("FY2025_Base") or cc.findtext("FY2025")
                )
                results.append({
                    "PENumber": _norm(pe_number),
                    "PEName": _norm(pe_name),
                    "ProjectNumber": _norm(project_number),
                    "ProjectName": _norm(project_name),
                    "CostCategory": _norm(cc_name),
                    "FY2023_Cost": fy23,
                    "FY2024_Cost": fy24,
                    "FY2025_Base_Cost": fy25_base,
                })
    return results


def _find_header_row(sheet, max_search_rows: int = 50) -> Optional[int]:
    current_row = 0
    for row in sheet.iter_rows(min_row=1, max_row=max_search_rows):
        current_row = row[0].row if row else current_row + 1
        values = [str(c.value).strip().lower() if c.value is not None else "" for c in row]
        joined = " ".join(values)
        if all(k in joined for k in ["project", "program", "element"]):
            return current_row
        # common case: headers in a single row
        if any("program element" in v for v in values) and any("project" in v for v in values):
            return current_row
        # R-1D style headers: Type, PE#, Project#, Title, Description, FY columns
        if any(v == "type" for v in values) and any("pe#" in v for v in values) and any("project#" in v for v in values):
            return current_row
    return None


def _header_map(cells: Iterable) -> Dict[str, int]:
    mapping: Dict[str, int] = {}
    for idx, cell in enumerate(cells):
        v = str(cell.value).strip().lower() if cell.value is not None else ""
        if not v:
            continue
        if "program element" in v and ("number" in v or v.endswith("program element")):
            mapping.setdefault("PENumber", idx)
        if ("program element name" in v) or ("pe name" in v) or (v.endswith("title") and "program element" in v):
            mapping.setdefault("PEName", idx)
        if ("project number" in v) or (v == "project") or (v == "project#"):
            mapping.setdefault("ProjectNumber", idx)
        if ("project name" in v) or ("project title" in v):
            mapping.setdefault("ProjectName", idx)
        if ("cost category" in v) or ("cost element" in v) or ("line item" in v):
            mapping.setdefault("CostCategory", idx)
        # R-1D variants
        if v == "pe#":
            mapping.setdefault("PENumber", idx)
        if v == "type":
            mapping.setdefault("Type", idx)
        if "accomplishments/planned programs title" in v or v.startswith("pe/project/"):
            mapping.setdefault("Title", idx)
        if v == "description":
            mapping.setdefault("Description", idx)
        # FY columns
        if "fy2023" in v or v == "fy 2023" or v == "fy23":
            mapping.setdefault("FY2023_Cost", idx)
        if "fy2024" in v or v == "fy 2024" or v == "fy24":
            mapping.setdefault("FY2024_Cost", idx)
        if "fy 2025 base" in v or v == "fy25 base" or ("fy2025" in v and "base" in v) or ("fy2025 base" in v):
            mapping.setdefault("FY2025_Base_Cost", idx)
    return mapping


def _parse_r1d_sheet(sheet, header_row: int, hdr_map: Dict[str, int]) -> List[Dict]:
    results: List[Dict] = []
    current_pe_num: Optional[str] = None
    current_pe_name: Optional[str] = None
    current_proj_num: Optional[str] = None
    current_proj_name: Optional[str] = None

    def gv(row_vals, key: str) -> Optional[str]:
        idx = hdr_map.get(key)
        if idx is None or idx >= len(row_vals):
            return None
        v = row_vals[idx]
        return str(v).strip() if v is not None else None

    # iterate rows
    blank_streak = 0
    for cells in sheet.iter_rows(min_row=header_row + 1, max_row=sheet.max_row):
        row_vals = [c.value for c in cells]
        if all(c is None or (isinstance(c, str) and not c.strip()) for c in row_vals):
            blank_streak += 1
            if blank_streak >= 10:
                break
            continue
        blank_streak = 0

        typ = (gv(row_vals, "Type") or "").upper()
        pe_num = gv(row_vals, "PENumber") or current_pe_num
        proj_num = gv(row_vals, "ProjectNumber") or current_proj_num
        title = gv(row_vals, "Title") or ""

        # Update context
        if typ == "PE":
            current_pe_num = gv(row_vals, "PENumber") or current_pe_num
            current_pe_name = title or current_pe_name
            current_proj_num = None
            current_proj_name = None
            continue
        if typ == "PROJECT" or typ == "PROJECTS" or typ == "PRJ":
            current_proj_num = gv(row_vals, "ProjectNumber") or current_proj_num
            current_proj_name = title or current_proj_name
            # emit a project-level total row using FY columns if present
            rec = {
                "PENumber": _norm(current_pe_num),
                "PEName": _norm(current_pe_name),
                "ProjectNumber": _norm(current_proj_num),
                "ProjectName": _norm(current_proj_name),
                "CostCategory": "Project Totals",
                "FY2023_Cost": _parse_float_safe_any(gv(row_vals, "FY2023_Cost")),
                "FY2024_Cost": _parse_float_safe_any(gv(row_vals, "FY2024_Cost")),
                "FY2025_Base_Cost": _parse_float_safe_any(gv(row_vals, "FY2025_Base_Cost")),
                "R1D_Description": gv(row_vals, "Description") or "",
            }
            results.append(rec)
            continue

        # A/PP and CA are cost categories under current project context
        if typ in ("A/PP", "CA") or (typ and typ not in ("PE", "PROJECT")):
            rec = {
                "PENumber": _norm(current_pe_num),
                "PEName": _norm(current_pe_name),
                "ProjectNumber": _norm(current_proj_num),
                "ProjectName": _norm(current_proj_name),
                "CostCategory": _norm(title or gv(row_vals, "Description") or typ),
                "FY2023_Cost": _parse_float_safe_any(gv(row_vals, "FY2023_Cost")),
                "FY2024_Cost": _parse_float_safe_any(gv(row_vals, "FY2024_Cost")),
                "FY2025_Base_Cost": _parse_float_safe_any(gv(row_vals, "FY2025_Base_Cost")),
                "R1D_Description": gv(row_vals, "Description") or "",
            }
            # Skip totals summary rows to avoid double-counting
            cc_lower = rec["CostCategory"].lower()
            if ("totals" in cc_lower and "sum of" in cc_lower) or cc_lower in ("project totals",):
                continue
            # Only add if there is some identifier
            if rec["PENumber"] or rec["ProjectNumber"] or rec["CostCategory"]:
                results.append(rec)
            continue

    return results


def _parse_xlsx_file(path: str) -> List[Dict]:
    results: List[Dict] = []
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return results
    for sheet in wb.worksheets:
        header_row = _find_header_row(sheet)
        if not header_row:
            continue
        # Retrieve header cells safely via iter_rows
        try:
            header_cells = next(sheet.iter_rows(min_row=header_row, max_row=header_row))
        except StopIteration:
            continue
        hdr_map = _header_map(header_cells)
        if not hdr_map:
            continue
        # If this looks like R-1D (Type + PE# + Project#), use specialized parsing
        if {"Type", "PENumber", "ProjectNumber"}.issubset(hdr_map.keys()):
            results.extend(_parse_r1d_sheet(sheet, header_row, hdr_map))
            continue
        # iterate data rows until we hit many blanks in a row
        blank_streak = 0
        for cells in sheet.iter_rows(min_row=header_row + 1, max_row=sheet.max_row):
            row_vals = [c.value for c in cells]
            if all(c is None or (isinstance(c, str) and not c.strip()) for c in row_vals):
                blank_streak += 1
                if blank_streak >= 5:
                    break
                continue
            blank_streak = 0
            def gv(key: str) -> Optional[str]:
                idx = hdr_map.get(key)
                if idx is None or idx >= len(row_vals):
                    return None
                v = row_vals[idx]
                return str(v).strip() if v is not None else None
            rec = {
                "PENumber": _norm(gv("PENumber")),
                "PEName": _norm(gv("PEName")),
                "ProjectNumber": _norm(gv("ProjectNumber")),
                "ProjectName": _norm(gv("ProjectName")),
                "CostCategory": _norm(gv("CostCategory")),
                "FY2023_Cost": _parse_float_safe_any(gv("FY2023_Cost")),
                "FY2024_Cost": _parse_float_safe_any(gv("FY2024_Cost")),
                "FY2025_Base_Cost": _parse_float_safe_any(gv("FY2025_Base_Cost")),
            }
            # Skip rows with no PE or Project identifiers
            if not rec["PENumber"] and not rec["ProjectNumber"]:
                continue
            results.append(rec)
    return results


def parse_r3_projects(xml_paths: List[str]) -> List[Dict]:
    """
    Parse R-3 style project and cost item rows from embedded XML files.

    NOTE: The exact schema varies by service/volume. This function implements a
    conservative approach:
    - Scans for elements commonly used in R-3 exports (Program Element, Project, Cost lines)
    - Attempts to read per-FY amounts (FY2023, FY2024, FY2025 Base)

    Replace XPath selectors with concrete ones once an example XML is available.
    """
    results: List[Dict] = []

    for path in xml_paths:
        if not os.path.isfile(path):
            continue
        ext = os.path.splitext(path)[1].lower()
        if ext in (".xml", ".xls", ".xlt", ".mxl"):  # some Excel 2003 XML may have .xml
            results.extend(_parse_xml_file(path))
        elif ext in (".xlsx", ".xlsm"):
            results.extend(_parse_xlsx_file(path))
        else:
            # Try XML first, otherwise xlsx
            parsed = _parse_xml_file(path)
            if parsed:
                results.extend(parsed)
            else:
                results.extend(_parse_xlsx_file(path))

    return results
